import win32com.client as win32
import path
import xlrd
import csc.csclxmls
import openpyxl


# TODO: error handling , what ifs:
# TODO: attachment name exist but no file
# TODO: XML Corrupted, Requestor tag , if error go where and send to who . error log dialogue id.

# Drafting and sending email notification to senders. You can add other senders' email in the list
def send_outlookmail():
    """
     Name : send_outlookmail
     Description : To use outlook to send email.
     Prerequisites : outlook exe must be opened.
    """
    # getting xml path
    pth = path.path('XML\mail.xml')
    mxml = csc.csclxmls.node(pth)

    # create tables
    workbook = openpyxl.load_workbook('\EXCEL\output.xlsx')
    worksheet = workbook.get_sheet_by_name('Sheet1')
    body = '<p><table border="1">'
    row = 1

    while 1:
        col = 1
        appending = ""
        if worksheet.cell(row=row, column=col).value != None:
            appending += "<tr>"
            for x in range(0, 11):
                appending += "<td>" + str(worksheet.cell(row=row, column=col).value) + "</td>"
                col += 1
            body += appending + "</tr>"


        else:
            break
        row += 1
    body += "</table>"  # outlook create item
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    # mail.SentOnBehalfOfName = "mailer@company.com"
    mail.To = mxml('//item[@key="To"]').text
    mail.Cc = mxml('//item[@key="Cc"]').text
    mail.Bcc = mxml('//item[@key="Bcc"]').text
    mail.Subject = mxml('//item[@key="Subject"]').text
    # compile string for html body below
    mail.HTMLBody = "asd"
    # iterate each attachment

    mail.display(True)




    # Open Outlook.exe. Path may vary according to system config
    # Please check the path to .exe file and update below

    # def open_outlook():
    #     try:
    #         subprocess.call(['C:\Program Files\Microsoft Office\Office15\Outlook.exe'])
    #         os.system("C:\Program Files\Microsoft Office\Office15\Outlook.exe");
    #     except:
    #         print("Outlook didn't open successfully")
    #
    #
    # # Checking if outlook is already opened. If not, open Outlook.exe and send email
    # for item in psutil.pids():
    #     p = psutil.Process(item)
    #     if p.name() == "OUTLOOK.EXE":
    #         flag = 1
    #         break
    #     else:
    #         flag = 0
    #
    # if (flag == 1):
    #     send_outlookmail()
    # else:
    #     open_outlook()
    #     send_outlookmail()


send_outlookmail()
